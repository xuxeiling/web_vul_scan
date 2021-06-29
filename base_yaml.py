#实现yaml文件解析，支持yaml文件嵌套调用

from  openpyxl import load_workbook
from words import words
#先将excel 文件的内容原原本本的记录下来

def  getExcel(path = 'yongli.xlsx'):
    wb = load_workbook(path )
    # 步骤一：解析所有的用例步骤
    sh = wb[wb.sheetnames[0]]
    testinfo = {}
    steps = []
    action = False
    # print('sh.rows',sh.rows)
    for row in sh.rows:
        first_cell = row[0].value
        # print('firstcell',first_cell)
        if action is False:
            if first_cell is None:
                continue
            if len(row) < 2:
                continue
            if first_cell in [words.id, words.title, words.info]:
                # print('row[1].value',row[1].value)
                testinfo[first_cell] = row[1].value if row[1].value is not None else ''
            if first_cell in [words.weight]:
                try:
                    testinfo[first_cell] = int(row[1].value)
                except:
                    testinfo[first_cell] = 0
            if first_cell == words.module_type:
                action = True
        else:
            if first_cell is None:
                break
            ary_list = [words.module_type,
                        words.pay_load,
                        words.actinfo,
                       ]
            step = {}
            for index in range(len(row)):
                if index < len(ary_list):
                    cell_value = row[index].value
                    if cell_value is None:
                        continue
                    else:
                        step[ary_list[index]] = cell_value
                else:
                    break
            steps.append(step)
    if words.weight not in testinfo.keys():
        testinfo[words.weight] = 0

    # 步骤二 解析所有的参数化结果
    params = {}
    domains = []
    action2 = False
    if len(wb.sheetnames) > 1:
        sh2 = wb[wb.sheetnames[1]]
        for row in sh2.rows:
            first_cell = row[0].value
            if first_cell is None:
                continue
            # 当开始计数的时候，最后总要加一行 continue 以免误导后边的逻辑
            if action2:
                for index in range(1, len(row)):
                    if index > len(domains):
                        continue
                    cell_value = row[index].value
                    if cell_value is None or str(cell_value).strip(' ') == '':
                        continue
                    domain = domains[index - 1]
                    if domain not in params.keys():
                        params[domain] = {}
                    params[domain][str(first_cell).strip(' ')] = str(cell_value).strip(' ')
                continue
            if first_cell == words.params:
                action2 = True
                for index in range(1, len(row)):
                    cell_value = row[index].value
                    if cell_value is None:
                        break
                    cell_value = str(cell_value).strip(' ')
                    if cell_value == '':
                        raise Exception('domain can not be empty')
                    domains.append(cell_value)

    wb.close()
    # print("eeeeee",testinfo,steps)
    # return {
    #     words.testinfo: testinfo,  # 用例的关键信息
    #     words.steps: steps,  # 用例的步骤
    #     words.params: params,  # 用例的所有参数信息
    # }
    return testinfo,steps
if __name__ == '__main__':
    # print("cshi",type(getExcel(path='yongli.xlsx')))
    data = getExcel(path='yongli.xlsx')
    id = data[0].get('id')
    title = data[0].get('title')
    info = data[0].get('info')
    for i in data[1]:
        module = i.get('module_type')
        module2 = eval(i.get('pay_load'))
        # print('module:',module2)
        # print('payload:',eval(module2))
        if module == 'Strsql':
            # payload = eval(module2)
            # print(eval(module2)[0],eval(module2)[1],eval(module2)[2])
            print('payload:', module2[1])
            for j in  module2[1]:
                 print(j)
